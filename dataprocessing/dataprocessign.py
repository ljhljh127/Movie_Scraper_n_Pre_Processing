import os,sys
import pandas as pd
from openpyxl import load_workbook
path="C:/Users/이정현/PycharmProjects/pythonProject/dataprocessing/영화"
file_list = os.listdir(path)
for file_name_raw in file_list:
    file_name="C:/Users/이정현/PycharmProjects/pythonProject/dataprocessing/영화/"+file_name_raw
    print(file_name)
    wb=load_workbook(file_name)
    ws=wb.active
    ws.unmerge_cells("A1:O1")
    ws.unmerge_cells("A2:O2")
    ws.unmerge_cells("A3:O3")
    wb.save(file_name)
    df = pd.read_excel(file_name,engine='openpyxl')
    df=df.drop(df.columns[[2,6,7,8,9,11,12,14]],axis=1)
    df=df.drop(df.index[[0,1,2]],axis=0)
    df.to_excel(file_name)
    df.columns = ['날짜', '스크린수','상영횟수','상영점유율','좌석수','관객수','누적관객수']
    df.set_index('날짜',inplace=True)
    df.to_excel(file_name)



